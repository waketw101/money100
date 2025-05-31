import os
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import openpyxl
# 取得昨天的日期
yesterday = datetime.now() - timedelta(days=1)
date_str = yesterday.strftime("%Y%m%d")
#target_date = date_str
target_date = "20250529"
# 資料夾與檔案設定
folder_path = r"C:\Users\lanst\Desktop\股市\前進百分百"
output_file = fr"C:\Users\lanst\Desktop\股市\前進百分百\merged_{target_date}.txt"
csv_prefix = f"{target_date}"
csv_files = [f for f in os.listdir(folder_path)
             if f.endswith(".csv") and f.startswith(csv_prefix)]
xlsx_file = fr"C:\Users\lanst\Desktop\股市\排行榜標的\{target_date}_成量排行.xlsx"  # today_str=target_date

if not os.path.exists(output_file):
    # 找出所有包含日期字串的 txt 檔
    txt_files = [f for f in os.listdir(folder_path)
                 if f.endswith(".txt") and target_date in f]
    txt_files = txt_files[:3]
    # 合併檔案內容，並移除亂碼
    with open(output_file, "w", encoding="utf-8") as outfile:
        for fname in txt_files:
            with open(os.path.join(folder_path, fname), "r", encoding="utf-8") as infile:
                content = infile.read()
                # 移除常見亂碼，例如「��」及「�」
                content = content.replace("��", "").replace("�", "").replace(" ", "")
                outfile.write(content)
                outfile.write("\n")
    print(f"已合併 {len(txt_files)} 個檔案為 {output_file}")
else:
    # 2. 讀取 output 檔案內容，存成 set（方便比對）
    with open(output_file, "r", encoding="utf-8") as f:
        output_lines = set(line.strip() for line in f if line.strip())

    # 3. 讀取 xlsx 檔案，取得「代號」和「名稱」欄位
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    ws = wb.active

    # 假設第一列是標題，找出「代號」和「名稱」的欄位索引
    header = [cell.value for cell in ws[1]]
    code_idx = header.index("代號")
    name_idx = header.index("名稱")
    matched_codes = []

    # 4. 比對
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[code_idx]).strip()
        name = str(row[name_idx]).strip()
        # 判斷 code 或 name 是否出現在 output_lines
        if code in output_lines or name in output_lines:
            matched_codes.append(code)

    # 5. 輸出結果
    print(matched_codes)

