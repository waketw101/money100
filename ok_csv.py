import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# today_str = datetime.today().strftime('%Y%m%d')
today_str = '20250529'

# 讀取資料
excel_df = pd.read_csv(f"excel/csv2/{today_str}_成量排行2.csv", encoding='utf-8-sig')
csv_df = pd.read_csv(fr"C:\Users\lanst\Desktop\股市\國台資料\{today_str}.csv")
csv_df.rename(columns={"代碼": "代號"}, inplace=True)

# 指定要替換的欄位
columns_to_replace = ['均價', '昨量', '融資餘額', '融券餘額']
key_column = '代號'
excel_df.set_index(key_column, inplace=True)
csv_df.set_index(key_column, inplace=True)

for col in columns_to_replace:
    if col in csv_df.columns and col in excel_df.columns:
        excel_df[col] = csv_df[col]

if '換手率%' in excel_df.columns:
    excel_df.drop(columns=['換手率%'], inplace=True)
if '換手率%' in csv_df.columns:
    excel_df.insert(6, '換手率%', csv_df['換手率%'])
if '成交' in excel_df.columns and '成本價' in excel_df.columns:
    excel_df['成本比'] = (excel_df['成交'] / excel_df['成本價']).round(2)
else:
    print("缺少 '成交' 或 '成本價' 欄位，無法計算成本比")

desired_order = [
    "代號", "名稱", "產業", "成交", "開盤", "最高", "最低", "漲跌幅", "成交張數", "昨量", "換手率%", "均價", "成本價",
    "成本比", "融資餘額", "融券餘額", "外資評分", "外資持有", "外資近三月", "營收評分", "近期營收", "近期esp", "本益比",
    "近三營收", "近三季esp", "stock_id", "name"
]
cols_in_both = [col for col in desired_order if col in excel_df.columns]
excel_df = excel_df[cols_in_both + [col for col in excel_df.columns if col not in cols_in_both]]
excel_df.reset_index(inplace=True)
output_path = fr"C:\Users\lanst\Desktop\股市\排行榜標的\{today_str}_成量排行.xlsx"
excel_df.to_excel(output_path, index=False)

# --- Excel 樣式處理 ---
wb = load_workbook(output_path)
ws = wb.active

header = [cell.value for cell in ws[1]]
col_idx = {col: idx+1 for idx, col in enumerate(header)}  # Excel從1開始

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # 預設每格字型
    for cell in row:
        cell.font = Font(size=12)
    # 漲跌幅 > 6，紅色粗體
    try:
        v = float(row[col_idx['漲跌幅']-1].value or 0)
        if v > 6:
            row[col_idx['漲跌幅']-1].font = Font(size=12, color="C00000", bold=True)
        elif v < 0:
            row[col_idx['漲跌幅']-1].font = Font(size=12, color="00B000", bold=True)
    except Exception:
        pass
    # 成交張數 >= 10000，紅色粗體
    try:
        v = float(row[col_idx['成交張數']-1].value or 0)
        if v >= 10000:
            row[col_idx['成交張數']-1].font = Font(size=12, color="C00000", bold=True)
    except Exception:
        pass
    # 成本比 >= 1.13 紅色粗體, <0.9 綠色粗體
    try:
        v = float(row[col_idx['成本比']-1].value or 0)
        if v >= 1.13:
            row[col_idx['成本比']-1].font = Font(size=12, color="C00000", bold=True)
        elif v < 0.9:
            row[col_idx['成本比']-1].font = Font(size=12, color="00B000", bold=True)
    except Exception:
        pass
    # 營收評分 >= 3 紅色粗體, <0 綠色粗體
    try:
        v = float(row[col_idx['營收評分']-1].value or 0)
        if v >= 3:
            row[col_idx['營收評分']-1].font = Font(size=12, color="C00000", bold=True)
        elif v < 0:
            row[col_idx['營收評分']-1].font = Font(size=12, color="00B000", bold=True)
    except Exception:
        pass
    # 近期營收 >= 50 紅色粗體, <0 綠色粗體
    try:
        cell_val = row[col_idx['近期營收']-1].value
        if isinstance(cell_val, str) and '%' in cell_val:
            v = float(cell_val.replace('%',''))
        else:
            v = float(cell_val or 0)
        if v >= 50:
            row[col_idx['近期營收']-1].font = Font(size=12, color="C00000", bold=True)
        elif v < 0:
            row[col_idx['近期營收']-1].font = Font(size=12, color="00B000", bold=True)
    except Exception:
        pass

# 標題列字型
for cell in ws[1]:
    cell.font = Font(size=12, bold=True)

wb.save(output_path)